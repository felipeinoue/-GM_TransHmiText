from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import *

class Text:
    def __init__(self, english, portuguese):
        self.english = english
        self.portuguese = portuguese

def TranslateHmiText(filename):

    # load files
    fileRef = 'CM010.xlsx'
    fileDest = 'projeto base.xlsx'

    # get reference workbook
    try:
        wb_ref = load_workbook(fileRef)
    except:
        pass
        return {
            'message': "Couldn't open selected file"
        }

    # get worksheet
    ws_ref = wb_ref['User Texts']

    # build list of Text reference objects
    ls_ObjectsRef = []
    x = 0
    for x in range(ws_ref.max_row):

        # get row
        row = x+1

        # build text object
        text = Text(
            english = ws_ref[f'E{row}'].value,
            portuguese = ws_ref[f'F{row}'].value
        )

        # append to list
        ls_ObjectsRef.append(text)


    # get destination workbook
    try:
        wb_dest = load_workbook(fileDest)
    except:
        pass
        return {
            'message': "Couldn't open selected file"
        }

    # get destination
    ws_dest = wb_dest['User Texts']

    # build list of Text destination objects
    ls_ObjectsDest = []
    x = 0
    for x in range(ws_dest.max_row):

        # get row
        row = x+1

        # build text object
        text = Text(
            english = ws_dest[f'E{row}'].value,
            portuguese = ws_dest[f'F{row}'].value
        )

        # append to list
        ls_ObjectsDest.append(text)

    # build translated list
    ls_trans = []
    for ObjectDest in ls_ObjectsDest:

        # start variables
        found = False

        # get object ref
        for ObjectRef in ls_ObjectsRef:


            # check if english text is equal
            if ObjectDest.english == ObjectRef.english:
                ls_trans.append([ObjectDest.english, ObjectRef.portuguese])
                found = True
                break

        # translation not found
        if not found:
            ls_trans.append([ObjectDest.english, 'TRANSLATION NOT FOUND'])
            found = False


    # fim
    ls_trans.append(['end', 'fim'])

    # save to file
    wb = Workbook()

    # grab the active worksheet
    ws = wb.active

    # Rows can also be appended
    for row in ls_trans:
        ws.append(row)

    # get current time
    now = datetime.now()
    now_txt = now.strftime("_%Y%m%d_%H%M%S")

    # Save the file
    wb.save(f"export{now_txt}.xlsx")

    return {
        'message': 'Exported alarms succesfully!'
    }

TranslateHmiText('xx')
